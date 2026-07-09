"""
excel_writer.py — Generic Excel dumper. No fixed template.

Claude decides the sheets and columns; this module just writes them out:
bold blue header row, bordered cells, right-aligned numbers (negatives in
red), auto-fitted column widths, sanitised + deduplicated sheet names.
"""
from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_BG = "1F4E78"
BORDER_GREY = "BFBFBF"
NUM_FMT = '#,##0.00;[Red]-#,##0.00;"-"'


def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)


def _thin():
    s = Side(style="thin", color=BORDER_GREY)
    return Border(left=s, right=s, top=s, bottom=s)


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _safe_sheet_name(name, used):
    """Excel sheet names: <=31 chars, no []:*?/\\ characters, unique."""
    name = re.sub(r"[\[\]:*?/\\]", " ", str(name or "Sheet")).strip() or "Sheet"
    name = name[:31]
    base, i = name, 1
    while name.lower() in used:
        suffix = f" ({i})"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def _write_sheet(wb, sheet, used_names):
    ws = wb.create_sheet(_safe_sheet_name(sheet.get("name"), used_names))
    ws.sheet_view.showGridLines = False

    columns = [str(c) for c in (sheet.get("columns") or [])]
    rows = sheet.get("rows") or []

    if columns:
        for ci, h in enumerate(columns, 1):
            c = ws.cell(1, ci, h)
            c.font = Font(name="Calibri", bold=True, color="FFFFFF")
            c.fill = _fill(HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _thin()
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        start = 2
    else:
        start = 1

    n_cols = len(columns)
    for ri, row in enumerate(rows, start):
        cells = list(row) if isinstance(row, (list, tuple)) else [row]
        for ci, v in enumerate(cells, 1):
            if not isinstance(v, (int, float, str)) and v is not None:
                v = str(v)  # guard against nested lists/objects from the model
            c = ws.cell(ri, ci, v)
            c.border = _thin()
            if _is_num(v):
                c.number_format = NUM_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(name="Calibri", color="9C0006" if v < 0 else "000000")
            else:
                c.font = Font(name="Calibri")
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        n_cols = max(n_cols, len(cells))

    # Column widths — fit content, capped so long text wraps instead
    for ci in range(1, n_cols + 1):
        longest = 0
        for r in ws.iter_rows(min_col=ci, max_col=ci, values_only=True):
            if r[0] is not None:
                longest = max(longest, len(str(r[0])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(longest + 3, 12), 55)


def workbook_to_excel(workbook: dict, report_text: str = "") -> bytes:
    """Build .xlsx bytes from Claude's freeform {"sheets": [...]} structure.

    If no usable sheets came back, fall back to dumping the report text so
    the download is never empty.
    """
    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()
    for sheet in (workbook or {}).get("sheets") or []:
        if isinstance(sheet, dict):
            _write_sheet(wb, sheet, used_names)

    if not wb.sheetnames:
        ws = wb.create_sheet("Reconciliation Report")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 120
        for ri, line in enumerate((report_text or "No report was produced.").splitlines() or [""], 1):
            c = ws.cell(ri, 1, line)
            c.font = Font(name="Calibri")
            c.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
