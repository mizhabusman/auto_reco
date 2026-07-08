"""
writer.py — Build a neat, styled Excel report from Claude's structured result.

Fully generic: Claude decides the sections, columns and rows; this just renders
them into a clean workbook — a Summary sheet (overview + KPI metrics) followed
by one sheet per section, with coloured headers keyed to each section's "tone".
"""
from __future__ import annotations
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ─────────────────────────────────────────────────────────────────
INK        = "272220"
BRAND      = "8C5C2B"
BRAND_LT   = "F4EFE8"
BORDER_GREY = "D9D2C7"

TONE = {  # (header fill, header font, banner fill, banner font)
    "success": ("1E7D4F", "FFFFFF", "E5F4EC", "1E7D4F"),
    "warning": ("C77D14", "FFFFFF", "FBF1DC", "8A5A0F"),
    "danger":  ("B3261E", "FFFFFF", "F9E4E3", "B3261E"),
    "info":    ("1D6FA4", "FFFFFF", "E4F0F8", "1D6FA4"),
    "neutral": ("55504A", "FFFFFF", "F1EEE9", "55504A"),
}
NUM_FMT = '#,##0.00;[Red]-#,##0.00;"-"'


def _f(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)

def _font(size=11, bold=False, color=None, italic=False):
    kw = dict(name="Calibri", size=size, bold=bold, italic=italic)
    if color:
        kw["color"] = color
    return Font(**kw)

def _thin():
    s = Side(style="thin", color=BORDER_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _tone(name):
    return TONE.get(str(name or "").lower(), TONE["neutral"])

def _num(v):
    """Return a float if v looks numeric, else None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("₹", "").replace("$", "")
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            try:
                return float(s)
            except ValueError:
                return None
    return None

def _safe_sheet_name(name, used):
    name = re.sub(r'[\[\]\:\*\?\/\\]', " ", str(name or "Sheet")).strip() or "Sheet"
    name = name[:31]
    base, i = name, 1
    while name.lower() in used:
        suffix = f" ({i})"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name

def _autofit(ws, ncols, sample_rows):
    for ci in range(1, ncols + 1):
        letter = get_column_letter(ci)
        longest = 10
        for row in sample_rows:
            if ci - 1 < len(row):
                longest = max(longest, len(str(row[ci - 1])))
        ws.column_dimensions[letter].width = min(max(longest + 3, 12), 55)


# ── Summary sheet ────────────────────────────────────────────────────────────
def _summary_sheet(wb, summary, metrics, insights):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60

    c = ws.cell(1, 1, "Reconciliation Report")
    c.font = _font(size=18, bold=True, color=BRAND)
    ws.row_dimensions[1].height = 26

    r = 3
    if summary:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        sc = ws.cell(r, 1, summary)
        sc.font = _font(size=11, color=INK)
        sc.alignment = _align(wrap=True, v="top")
        # rough height for wrapped text
        ws.row_dimensions[r].height = max(45, 15 * (len(summary) // 110 + 1))
        r += 2

    if metrics:
        h = ws.cell(r, 1, "Key figures")
        h.font = _font(size=13, bold=True, color=INK)
        r += 1
        for m in metrics:
            hf = _tone(m.get("tone"))[0]
            lc = ws.cell(r, 1, str(m.get("label", "")))
            lc.font = _font(bold=True)
            lc.border = _thin()
            lc.alignment = _align()
            vc = ws.cell(r, 2, m.get("value", ""))
            vc.font = _font(bold=True, color=hf)
            vc.fill = _f(_tone(m.get("tone"))[2])
            vc.border = _thin()
            vc.alignment = _align(h="right")
            r += 1
        r += 1

    if insights:
        h = ws.cell(r, 1, "Findings & Recommendations")
        h.font = _font(size=13, bold=True, color=INK)
        r += 1
        for it in insights:
            hf, _, bfill, bfont = _tone(it.get("tone"))
            title  = str(it.get("title", "") or "").strip()
            detail = str(it.get("detail", it.get("text", "")) or "").strip()
            tc = ws.cell(r, 1, title or "Note")
            tc.font = _font(bold=True, color=bfont)
            tc.fill = _f(bfill)
            tc.border = _thin()
            tc.alignment = _align(v="top", wrap=True)
            dc = ws.cell(r, 2, detail)
            dc.font = _font(color=INK)
            dc.border = _thin()
            dc.alignment = _align(v="top", wrap=True)
            ws.row_dimensions[r].height = max(30, 15 * (len(detail) // 70 + 1))
            r += 1


# ── One sheet per section ─────────────────────────────────────────────────────
def _section_sheet(wb, section, used_names):
    title = section.get("title", "Section")
    ws = wb.create_sheet(_safe_sheet_name(title, used_names))
    ws.sheet_view.showGridLines = False

    hfill, hfont, bfill, bfont = _tone(section.get("tone"))
    columns = section.get("columns", []) or []
    rows    = section.get("rows", []) or []
    # Size to the widest of (declared columns, any row) so no cell is lost.
    ncols   = max([1, len(columns)] + [len(r) for r in rows])
    columns = [str(c) for c in columns] + [f"Col {i + 1}" for i in range(len(columns), ncols)]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(1, 1, title)
    tc.font = _font(size=15, bold=True, color=INK)
    ws.row_dimensions[1].height = 22

    r = 2
    note = section.get("note")
    if note:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        nc = ws.cell(r, 1, str(note))
        nc.font = _font(size=11, bold=True, color=bfont)
        nc.fill = _f(bfill)
        nc.alignment = _align(wrap=True)
        ws.row_dimensions[r].height = 26
        r += 1

    if not columns and not rows:
        ws.cell(r, 1, "(No entries)").font = _font(italic=True, color="909090")
        return

    # Header row
    header_row = r
    for ci, col in enumerate(columns, 1):
        hc = ws.cell(header_row, ci, str(col))
        hc.font = _font(bold=True, color=hfont)
        hc.fill = _f(hfill)
        hc.alignment = _align(h="center", wrap=True)
        hc.border = _thin()
    ws.row_dimensions[header_row].height = 20
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    # Data rows
    dr = header_row + 1
    for row in rows:
        cells = list(row) + [""] * (ncols - len(row))
        for ci, v in enumerate(cells[:ncols], 1):
            cell = ws.cell(dr, ci)
            cell.border = _thin()
            n = _num(v)
            if n is not None:
                cell.value = n
                cell.number_format = NUM_FMT
                cell.alignment = _align(h="right")
                cell.font = _font(color="B3261E") if n < 0 else _font()
            else:
                cell.value = "" if v is None else str(v)
                cell.alignment = _align(wrap=True)
                cell.font = _font()
        dr += 1

    _autofit(ws, ncols, [columns] + [list(x) for x in rows[:40]])


# ── Entry point ───────────────────────────────────────────────────────────────
def reco_to_excel(summary: str, metrics: list, insights: list, sections: list) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _summary_sheet(wb, summary, metrics or [], insights or [])

    used = {"summary"}
    for section in (sections or []):
        _section_sheet(wb, section, used)

    if len(wb.sheetnames) == 0:  # safety: never save an empty workbook
        wb.create_sheet("Summary")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
